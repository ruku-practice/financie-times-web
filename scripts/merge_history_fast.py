import os
import sys
import json
import re
import requests
import openpyxl
from google.oauth2.service_account import Credentials
from google.auth.transport.requests import Request

def main():
    scope = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
    credentials_path = "/Users/kurokzhr/Library/CloudStorage/GoogleDrive-ruku.practice@gmail.com/マイドライブ/00_XXX_TIMES/00_CreateAutoTimes/100_FiNANCiE/writeinfo2spreadsheet-d08cec7b431b.json"

    if not os.path.exists(credentials_path):
        print(f"Credentials not found at {credentials_path}")
        sys.exit(1)

    print("Authenticating with Google Drive API...")
    credentials = Credentials.from_service_account_file(credentials_path, scopes=scope)
    credentials.refresh(Request())
    access_token = credentials.token

    keys = {
        "2026": "1bjBz634jh4xIALM8nJ6G9V9MINWincIdYSUYa3w8qHo",
        "2025": "1ObSqlX6WhrywERsdYnHi59058AIydgt_UlEyXA005uM",
        "2024": "1mkOJctOSAJRninTTskXBxbXOssYjMJR62bBWjbQUNck"
    }

    os.makedirs("data", exist_ok=True)

    # 1. 各スプレッドシートを Excel としてエクスポートダウンロード
    for year, file_id in keys.items():
        dest_path = f"data/{year}.xlsx"
        if os.path.exists(dest_path):
            print(f"Excel file for {year} already exists at {dest_path}, skipping download.")
            continue
            
        print(f"Exporting {year} spreadsheet as Excel via Drive API...")
        url = f"https://www.googleapis.com/drive/v3/files/{file_id}/export?mimeType=application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        headers = {"Authorization": f"Bearer {access_token}"}
        
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            with open(dest_path, 'wb') as f:
                f.write(response.content)
            print(f"Successfully downloaded {year} data to {dest_path}")
        else:
            print(f"Failed to download {year} data: {response.status_code} - {response.text}")
            sys.exit(1)

    # 2. 2026年 Excel からプロジェクト一覧（listシート）を取得
    print("Loading active project list from 2026 Excel...")
    wb_2026 = openpyxl.load_workbook("data/2026.xlsx", read_only=True)
    if "list" not in wb_2026.sheetnames:
        print("Error: 'list' sheet not found in 2026 Excel.")
        sys.exit(1)
        
    list_sheet = wb_2026["list"]

    projects = []
    # Row 1 is header
    for row in list_sheet.iter_rows(min_row=2, values_only=True):
        # Col B (index 1): slug
        # Col C (index 2): folder (sheet name)
        # Col E (index 4): logo URL
        # Col G (index 6): title
        if len(row) > 2 and row[1] and row[2]:
            slug = str(row[1]).strip()
            folder = str(row[2]).strip()
            name = str(row[6]).strip() if len(row) > 6 and row[6] else slug
            logo = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            projects.append({
                "slug": slug,
                "folder": folder,
                "name": name,
                "logo": logo
            })

    print(f"Active projects count: {len(projects)}")
    wb_2026.close()

    # 3. データの結合マージ
    history = {}
    for proj in projects:
        history[proj["folder"]] = {
            "slug": proj["slug"],
            "name": proj["name"],
            "logo": proj["logo"],
            "data": {}
        }

    for year in ["2024", "2025", "2026"]:
        excel_path = f"data/{year}.xlsx"
        print(f"Parsing {year} Excel data...")
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        for p_idx, proj in enumerate(projects):
            folder_name = proj["folder"]
            if folder_name not in wb.sheetnames:
                continue
                
            sheet = wb[folder_name]
            # 全セルの値を 2次元配列としてロード
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
                
            date_row = rows[1] # Row 2 (index 1) が日付
            
            # 日付パターン一致列をスキャン
            for col_idx, val in enumerate(date_row):
                if val is None:
                    continue
                val_str = str(val).strip()
                if re.match(r'^\d{8}$', val_str):
                    date_str = val_str
                    
                    # 安全に値を取得するヘルパー
                    def get_val(row_idx):
                        if row_idx < len(rows) and col_idx < len(rows[row_idx]):
                            v = rows[row_idx][col_idx]
                            return v if v is not None else ""
                        return ""

                    # 数値変換ヘルパー
                    def get_float(row_idx):
                        v = get_val(row_idx)
                        if v == "": return 0.0
                        try: return float(str(v).replace(",", ""))
                        except: return 0.0
                        
                    def get_int(row_idx):
                        v = get_val(row_idx)
                        if v == "": return 0
                        try: return int(float(str(v).replace(",", "")))
                        except: return 0

                    # volumeの取得とクレンジング
                    volume_raw = get_val(11) # Row 12 (index 11)
                    volume = 0.0
                    if volume_raw != "":
                        try:
                            volume = float(str(volume_raw).replace(",", ""))
                            if volume >= 100000000.0:
                                volume = volume / (10**18)
                        except:
                            pass

                    close_price = get_float(12)    # Row 13 (index 12)
                    stock = get_int(13)            # Row 14 (index 13)
                    marketCap = get_int(14)        # Row 15 (index 14)
                    volume_data = get_float(15)    # Row 16 (index 15)
                    num_member = get_int(16)        # Row 17 (index 16)
                    active_ranking = str(get_val(17)).strip() # Row 18 (index 17)
                    current_price = get_float(18)  # Row 19 (index 18)

                    # マージ
                    history[folder_name]["data"][date_str] = {
                        "volume": round(volume, 4),
                        "close_price": close_price,
                        "stock": stock,
                        "marketCap": marketCap,
                        "volume_data": round(volume_data, 4),
                        "num_member": num_member,
                        "active_ranking": active_ranking,
                        "current_price": current_price
                    }
                    
        wb.close()
        print(f"Finished parsing {year} Excel.")

    # 4. history.json 保存
    print("Writing data/history.json...")
    with open("data/history.json", 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)
    print("Successfully generated data/history.json")

    # 一時的なExcelファイルをクリーンアップ（容量削減のため）
    print("Cleaning up temporary Excel files...")
    for year in keys.keys():
        try:
            os.remove(f"data/{year}.xlsx")
        except:
            pass

if __name__ == "__main__":
    main()
